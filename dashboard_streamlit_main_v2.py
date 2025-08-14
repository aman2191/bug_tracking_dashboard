import pandas as pd
from datetime import datetime, timedelta
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, BarChart3D, Reference
from openpyxl.chart.label import DataLabelList

# Check for required packages
try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    st.warning("Plotly is not installed. Charts will be displayed using Streamlit's native charts instead.")
    import matplotlib.pyplot as plt

# ==== CONFIGURATION ====
st.set_page_config(layout="wide", page_title="Bug Tracking Dashboard")

# ==== FILE UPLOAD ====
st.sidebar.header("Upload Files")
file1 = st.sidebar.file_uploader("Upload CSV File (data.csv)", type=["csv"])
file2 = st.sidebar.file_uploader("Upload Excel File (Issues Log.xlsx)", type=["xlsx"])

if not file1 or not file2:
    st.warning("Please upload both files to proceed")
    st.stop()

# ==== DATA PROCESSING WITH ERROR HANDLING ====
@st.cache_data
def load_and_process_data(file1, file2):
    try:
        df1 = pd.read_csv(file1)
        df2 = pd.read_excel(file2)
        
        # Check if ID column exists
        id_column = "ID"
        if id_column not in df1.columns or id_column not in df2.columns:
            st.error(f"'{id_column}' column not found in one of the files")
            st.stop()
        
        # Merge data
        matched_df = pd.merge(df1, df2, on=id_column, how='inner')
        
        # Process data with error handling
        matched_df['Date Raised'] = pd.to_datetime(matched_df['Date Raised'], errors='coerce')
        matched_df['Date Closed'] = pd.to_datetime(matched_df['Date Closed'], errors='coerce')
        matched_df['Days to Close'] = (matched_df['Date Closed'] - matched_df['Date Raised']).dt.days
        
        # Handle Assignee column - check if column exists first
        if 'Assigned To' in matched_df.columns:
            matched_df['Assignee'] = matched_df['Assigned To'].str.extract(r'^([^<]+)')[0].str.strip()
        else:
            matched_df['Assignee'] = 'Unassigned'
            st.warning("'Assigned To' column not found - using 'Unassigned' for all items")
        
        return matched_df
    
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        st.stop()

df = load_and_process_data(file1, file2)

# ==== DOWNLOAD BUTTONS ====
def get_excel_download_link(df, filename="merged_data.xlsx"):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Merged Data')
    output.seek(0)
    st.sidebar.download_button(
        label="📥 Download Merged Data",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def generate_excel_report(df):
    # Data Preparation
    df['Date Raised'] = pd.to_datetime(df['Date Raised'])
    df['Date Closed'] = pd.to_datetime(df['Date Closed'])
    df['Days to Close'] = (df['Date Closed'] - df['Date Raised']).dt.days
    df['Assignee'] = df['Assigned To'].str.extract(r'^([^<]+)')[0].str.strip()

    # =========================
    # WORKBOOK & STYLES
    # =========================
    wb = Workbook()
    
    # =========================
    # SHEET 1: MERGED DATA
    # =========================
    ws = wb.active
    ws.title = "Merged Data"

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # =========================
    # SHEET 2: DASHBOARD
    # =========================
    ws = wb.create_sheet(title="Dashboard")

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Page title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Bug Tracking Dashboard"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = header_fill

    # Layout helpers
    table_start_col = 'A'
    chart_start_col = 'E'
    current_row = 5  # first section row

    # Helper functions
    def put_section_header(ws, cell, title):
        ws.merge_cells(f'{cell}:B{cell[1:]}')
        ws[cell] = title
        ws[cell].font = Font(size=14, bold=True, color="FFFFFF")
        ws[cell].fill = header_fill
        ws[cell].border = border

    def put_chart_heading(ws, anchor_cell, title, span_cols=8, rows_above=1):
        col_letters = ''.join(filter(str.isalpha, anchor_cell))
        row_number = int(''.join(filter(str.isdigit, anchor_cell)))
        heading_row = row_number - rows_above
        start_col_idx = ws[f"{col_letters}{heading_row}"].column
        end_col_letter = get_column_letter(start_col_idx + span_cols - 1)
        ws.merge_cells(f"{col_letters}{heading_row}:{end_col_letter}{heading_row}")
        hcell = ws[f"{col_letters}{heading_row}"]
        hcell.value = title
        hcell.font = Font(size=14, bold=True)
        hcell.alignment = Alignment(horizontal="left", vertical="center")

    # =========================
    # KEY METRICS TABLE
    # =========================
    put_section_header(ws, f'{table_start_col}{current_row}', "Key Metrics")

    metrics = [
        ("Total Items", len(df)),
        ("Open Items", len(df[df['State_x'].isin(['Active', 'New', 'Blocked'])])),
        ("Closed Items", len(df[df['State_x'] == 'Closed'])),
        ("Bugs Raised This Week",
         len(df[(df['Date Raised'] >= datetime.now() - timedelta(days=7)) &
                (df['Work Item Type'] == 'Bug')])),
        ("Bugs Closed This Week",
         len(df[(df['Date Closed'] >= datetime.now() - timedelta(days=7)) &
                (df['Work Item Type'] == 'Bug') &
                (df['State_x'] == 'Closed')])),
        ("Hotfixes Deployed This Week",
         len(df[(df['Date Closed'] >= datetime.now() - timedelta(days=7)) &
                (df['Tags'].str.contains('Hot Fix', na=False))])),
        ("Avg Time to Close (Closed Items)",
         round(df[df['State_x'] == 'Closed']['Days to Close'].mean(), 1) 
         if not df[df['State_x'] == 'Closed'].empty else 0),
        ("Avg Days Open Bugs Pending",
         round((datetime.now() - df[df['State_x'].isin(['Active', 'New', 'Blocked'])]['Date Raised']).dt.days.mean(), 1)
         if not df[df['State_x'].isin(['Active', 'New', 'Blocked'])].empty else 0),
        ("Oldest Open Bug (days)",
         (datetime.now() - df[df['State_x'].isin(['Active', 'New', 'Blocked'])]['Date Raised'].min()).days
         if not df[df['State_x'].isin(['Active', 'New', 'Blocked'])].empty else 0)
    ]

    # Write metrics table
    start_metrics_row = current_row + 1
    for r_offset, (metric, value) in enumerate(metrics, start=0):
        r = start_metrics_row + r_offset
        ws[f'A{r}'] = metric
        ws[f'B{r}'] = value
        for col in ['A', 'B']:
            ws[f'{col}{r}'].border = border

    # =========================
    # KEY METRICS CHART (Fixed Version)
    # =========================
    # Define all metrics you want to include
    chart_metrics = [
        "Total Items",
        "Open Items",
        "Closed Items",
        "Bugs Raised This Week",
        "Bugs Closed This Week",
        "Hotfixes Deployed This Week",
        "Avg Time to Close (Closed Items)",
        "Avg Days Open Bugs Pending",
        "Oldest Open Bug (days)"
    ]

    # Find starting row (assuming metrics start at row 6)
    start_row = 6
    end_row = start_row + len(chart_metrics) - 1

    # Create the chart
    metric_chart = BarChart3D()
    metric_chart.title = None
    metric_chart.y_axis.title = "Count/Days"
    metric_chart.style = 10
    # metric_chart.height = 15  # Make taller to accommodate more bars
    # metric_chart.width = 25   # Make wider for better label display

    # Add data and categories
    data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)  # Column B
    cats_ref = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)  # Column A

    metric_chart.add_data(data_ref, titles_from_data=True)
    metric_chart.set_categories(cats_ref)

    # Format data labels
    metric_chart.dLbls = DataLabelList()
    metric_chart.dLbls.showVal = True
    metric_chart.dLbls.font = Font(size=8, bold=True)

    # Position chart (adjust anchor position as needed)
    chart_anchor = "E5"  # Moved up to accommodate more rows
    ws.add_chart(metric_chart, chart_anchor)
    put_chart_heading(ws, chart_anchor, "Key Metrics Summary", span_cols=8, rows_above=1)
    current_row += len(metrics) + 13  # space after section

    # =========================
    # STATE DISTRIBUTION
    # =========================
    put_section_header(ws, f'{table_start_col}{current_row}', "State Distribution")

    # headers
    ws[f'A{current_row+1}'] = "State"; ws[f'A{current_row+1}'].font = Font(bold=True)
    ws[f'B{current_row+1}'] = "Count"; ws[f'B{current_row+1}'].font = Font(bold=True)
    for col in ['A', 'B']:
        ws[f'{col}{current_row+1}'].fill = light_blue_fill
        ws[f'{col}{current_row+1}'].border = border

    state_counts = df['State_x'].value_counts().reset_index()
    state_counts.columns = ['State', 'Count']
    for i, (state, count) in enumerate(state_counts.values, start=current_row+2):
        ws[f'A{i}'] = state
        ws[f'B{i}'] = count
        ws[f'A{i}'].border = border; ws[f'B{i}'].border = border

    pie = PieChart()
    pie.title = None  # heading on sheet instead
    labels = Reference(ws, min_col=1, min_row=current_row+2, max_row=current_row+1+len(state_counts))
    data = Reference(ws, min_col=2, min_row=current_row+1, max_row=current_row+1+len(state_counts))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    state_chart_anchor = f"{chart_start_col}{current_row}"
    ws.add_chart(pie, state_chart_anchor)
    put_chart_heading(ws, state_chart_anchor, "State Distribution", span_cols=8, rows_above=1)

    current_row += len(state_counts) + 13

    # =========================
    # SPRINT ASSIGNMENT
    # =========================
    put_section_header(ws, f'{table_start_col}{current_row}', "Sprint Assignment")

    ws[f'A{current_row+1}'] = "Sprint"; ws[f'A{current_row+1}'].font = Font(bold=True)
    ws[f'B{current_row+1}'] = "Count"; ws[f'B{current_row+1}'].font = Font(bold=True)
    for col in ['A', 'B']:
        ws[f'{col}{current_row+1}'].fill = light_blue_fill
        ws[f'{col}{current_row+1}'].border = border

    sprint_counts = df['State_y'].value_counts().reset_index()
    sprint_counts.columns = ['Sprint', 'Count']
    for i, (sprint, count) in enumerate(sprint_counts.values, start=current_row+2):
        ws[f'A{i}'] = sprint
        ws[f'B{i}'] = count
        ws[f'A{i}'].border = border; ws[f'B{i}'].border = border

    bar = BarChart3D()
    bar.title = None
    bar.y_axis.title = "Count"
    bar.style = 10
    bar.dLbls = DataLabelList(showVal=True, showLegendKey=False,
                            showCatName=False, showSerName=False, showPercent=False)

    labels = Reference(ws, min_col=1, min_row=current_row+2, max_row=current_row+1+len(sprint_counts))
    data = Reference(ws, min_col=2, min_row=current_row+1, max_row=current_row+1+len(sprint_counts))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)

    sprint_chart_anchor = f"{chart_start_col}{current_row}"
    ws.add_chart(bar, sprint_chart_anchor)
    put_chart_heading(ws, sprint_chart_anchor, "Sprint Assignment", span_cols=8, rows_above=1)

    current_row += len(sprint_counts) + 13

    # =========================
    # ASSIGNEE WORKLOAD
    # =========================
    put_section_header(ws, f'{table_start_col}{current_row}', "Assignee Workload")

    ws[f'A{current_row+1}'] = "Assignee"; ws[f'A{current_row+1}'].font = Font(bold=True)
    ws[f'B{current_row+1}'] = "Task Count"; ws[f'B{current_row+1}'].font = Font(bold=True)
    for col in ['A', 'B']:
        ws[f'{col}{current_row+1}'].fill = light_blue_fill
        ws[f'{col}{current_row+1}'].border = border

    assignee_counts = df['Assignee'].fillna("Unassigned").value_counts().reset_index()
    assignee_counts.columns = ['Assignee', 'Task Count']
    for i, (assignee, count) in enumerate(assignee_counts.values, start=current_row+2):
        ws[f'A{i}'] = assignee
        ws[f'B{i}'] = count
        ws[f'A{i}'].border = border; ws[f'B{i}'].border = border

    assignee_bar = BarChart3D()
    assignee_bar.title = None
    assignee_bar.y_axis.title = "Task Count"
    assignee_bar.style = 10
    assignee_bar.dLbls = DataLabelList(showVal=True, showLegendKey=False,
                                    showCatName=False, showSerName=False, showPercent=False)

    labels = Reference(ws, min_col=1, min_row=current_row+2, max_row=current_row+1+len(assignee_counts))
    data = Reference(ws, min_col=2, min_row=current_row+1, max_row=current_row+1+len(assignee_counts))
    assignee_bar.add_data(data, titles_from_data=True)
    assignee_bar.set_categories(labels)

    assignee_chart_anchor = f"{chart_start_col}{current_row}"
    ws.add_chart(assignee_bar, assignee_chart_anchor)
    put_chart_heading(ws, assignee_chart_anchor, "Assignee Workload", span_cols=8, rows_above=1)

    # Adjust column widths if needed
    # ws.column_dimensions['A'].width = 30  # Wider for long metric names
    # ws.column_dimensions['B'].width = 15
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.freeze_panes = 'A4'
    

    return wb

def get_report_download_link(wb, filename="dashboard_report.xlsx"):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.sidebar.download_button(
        label="📊 Download Dashboard Report",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Display download buttons in sidebar
get_excel_download_link(df)

if st.sidebar.button("🖨️ Generate Excel Report"):
    with st.spinner("Generating Excel report..."):
        report_wb = generate_excel_report(df)
        get_report_download_link(report_wb)
    st.sidebar.success("Excel report generated!")

# ==== HELPER FUNCTIONS ====
def display_metric(label, value, suffix=""):
    if isinstance(value, float):
        value = round(value, 2)
    st.metric(label, f"{value}{suffix}")

def create_bar_chart(data, x, y, title):
    if PLOTLY_AVAILABLE:
        fig = px.bar(data, x=x, y=y, title=title, text=y)
        fig.update_traces(textposition='outside', marker_color='#4F81BD')
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.bar_chart(data.set_index(x)[y])

def create_pie_chart(data, names, values, title):
    if PLOTLY_AVAILABLE:
        fig = px.pie(data, names=names, values=values, title=title)
        st.plotly_chart(fig, use_container_width=True)
    else:
        fig, ax = plt.subplots()
        ax.pie(data[values], labels=data[names], autopct='%1.1f%%')
        ax.set_title(title)
        st.pyplot(fig)

# ==== DASHBOARD LAYOUT ====
st.title("Bug Tracking Dashboard")

# ==== KEY METRICS ====
st.header("Key Metrics")
col1, col2, col3, col4 = st.columns(4)

with col1:
    display_metric("Total Items", len(df))
    display_metric("Open Items", len(df[df['State_x'].isin(['Active', 'New', 'Blocked'])]))
    
with col2:
    display_metric("Closed Items", len(df[df['State_x'] == 'Closed']))
    display_metric("Bugs Raised This Week", 
                  len(df[(df['Date Raised'] >= datetime.now() - timedelta(days=7)) &
                         (df['Work Item Type'] == 'Bug')]))
    
with col3:
    display_metric("Bugs Closed This Week",
                  len(df[(df['Date Closed'] >= datetime.now() - timedelta(days=7)) &
                         (df['Work Item Type'] == 'Bug') &
                         (df['State_x'] == 'Closed')]))
    display_metric("Hotfixes Deployed This Week",
                  len(df[(df['Date Closed'] >= datetime.now() - timedelta(days=7)) &
                         (df['Tags'].str.contains('Hot Fix', na=False))]))
    
with col4:
    avg_close = (float(round(df[df['State_x'] == 'Closed']['Days to Close'].mean(), 1))
                 if not df[df['State_x'] == 'Closed'].empty else 0.0)
    display_metric("Avg Time to Close", avg_close, " days")
    
    avg_open = (float(round((datetime.now() - df[df['State_x'].isin(['Active', 'New', 'Blocked'])]['Date Raised']).dt.days.mean(), 1))
                if not df[df['State_x'].isin(['Active', 'New', 'Blocked'])].empty else 0.0)
    display_metric("Avg Days Open Bugs Pending", avg_open, " days")

# Key Metrics Chart
chart_metrics_data = {
    "Metric": ["Open Items", "Closed Items", "Bugs This Week", "Hotfixes"],
    "Count": [
        len(df[df['State_x'].isin(['Active', 'New', 'Blocked'])]),
        len(df[df['State_x'] == 'Closed']),
        len(df[(df['Date Raised'] >= datetime.now() - timedelta(days=7)) &
               (df['Work Item Type'] == 'Bug')]),
        len(df[(df['Date Closed'] >= datetime.now() - timedelta(days=7)) &
               (df['Tags'].str.contains('Hot Fix', na=False))])
    ]
}
create_bar_chart(pd.DataFrame(chart_metrics_data), "Metric", "Count", "Key Metrics Summary")

# ==== STATE DISTRIBUTION ====
st.header("State Distribution")
col1, col2 = st.columns([1, 2])

with col1:
    state_counts = df['State_x'].value_counts(dropna=False).reset_index()
    state_counts.columns = ['State', 'Count']
    state_counts['State'] = state_counts['State'].fillna("Unspecified")
    st.dataframe(state_counts, hide_index=True)

with col2:
    create_pie_chart(state_counts, 'State', 'Count', "State Distribution")

# ==== SPRINT ASSIGNMENT ====
st.header("Sprint Assignment")
col1, col2 = st.columns([1, 2])

with col1:
    sprint_counts = df['State_y'].value_counts(dropna=False).reset_index()
    sprint_counts.columns = ['Sprint', 'Count']
    sprint_counts['Sprint'] = sprint_counts['Sprint'].fillna("Unspecified")
    st.dataframe(sprint_counts, hide_index=True)

with col2:
    create_bar_chart(sprint_counts, 'Sprint', 'Count', "Sprint Assignment")

# ==== ASSIGNEE WORKLOAD ====
st.header("Assignee Workload")
col1, col2 = st.columns([1, 2])

with col1:
    assignee_counts = df['Assignee'].value_counts().reset_index()
    assignee_counts.columns = ['Assignee', 'Task Count']
    st.dataframe(assignee_counts, hide_index=True)

with col2:
    create_bar_chart(assignee_counts, 'Assignee', 'Task Count', "Assignee Workload")

# ==== ADDITIONAL VISUALIZATIONS ====
st.header("Additional Insights")

# Time to Close Distribution
if not df[df['State_x'] == 'Closed'].empty:
    closed_items = df[df['State_x'] == 'Closed']
    if PLOTLY_AVAILABLE:
        fig = px.histogram(closed_items, x='Days to Close', 
                         title="Distribution of Time to Close (Closed Items)")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.bar_chart(closed_items['Days to Close'].value_counts())

# Work Item Type Distribution
if 'Work Item Type' in df.columns:
    type_counts = df['Work Item Type'].value_counts().reset_index()
    type_counts.columns = ['Type', 'Count']
    create_pie_chart(type_counts, 'Type', 'Count', "Work Item Type Distribution")
else:
    st.warning("'Work Item Type' column not found - skipping Work Item Type chart")